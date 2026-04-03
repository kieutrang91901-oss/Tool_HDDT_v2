"""
ExcelExportService — Export danh sách & chi tiết hóa đơn ra Excel.

Tuân thủ:
- Cột export = cột user đã chọn hiển thị (ColumnConfig)
- Các cột tiền/số lượng tự động format #,##0
- Hỗ trợ cả cột Tầng 1 (fixed) và Tầng 2 (dynamic extras)
"""
import os
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from app.models.entities import InvoiceData, ColumnConfig
from config.column_config import parse_dynamic_column_key
from config.logger import get_logger

logger = get_logger(__name__)


# ═══════════════════════════════════════════════════════
# EXCEL STYLES
# ═══════════════════════════════════════════════════════

HEADER_FONT = Font(name="Inter", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Inter", size=11)
BODY_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
BODY_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
BODY_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

TITLE_FONT = Font(name="Inter", size=14, bold=True, color="6C63FF")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# Number format cho VNĐ (không thập phân)
VND_FORMAT = '#,##0'
NUMBER_FORMAT = '#,##0'
DECIMAL_FORMAT = '#,##0.00'


class ExcelExportService:
    """Export hóa đơn ra file Excel."""
    
    # ═══════════════════════════════════════════════════════
    # EXPORT SUMMARY (Bảng tổng hợp)
    # ═══════════════════════════════════════════════════════
    
    @staticmethod
    def export_summary(
        invoices: List[InvoiceData],
        dest_path: str,
        visible_columns: List[ColumnConfig],
    ) -> dict:
        """Export bảng tổng hợp theo cột user đã chọn.
        
        Returns:
            {"success": bool, "file_path": str, "row_count": int, "error_msg": str}
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tổng hợp HĐ"
            
            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(visible_columns))
            title_cell = ws.cell(row=1, column=1, value="DANH SÁCH HÓA ĐƠN ĐIỆN TỬ")
            title_cell.font = TITLE_FONT
            title_cell.alignment = Alignment(horizontal="center")
            
            # Headers (row 3)
            header_row = 3
            for col_idx, col_cfg in enumerate(visible_columns, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=col_cfg.display_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    col_cfg.width / 8, 12
                )
            
            # Data rows
            for row_idx, inv in enumerate(invoices, header_row + 1):
                for col_idx, col_cfg in enumerate(visible_columns, 1):
                    value = ExcelExportService._get_summary_value(inv, col_cfg, row_idx - header_row)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER
                    
                    # Apply format
                    ExcelExportService._apply_format(cell, col_cfg.format_type)
            
            # Auto-filter
            if invoices:
                ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(visible_columns))}{header_row + len(invoices)}"
            
            # Save
            os.makedirs(os.path.dirname(dest_path) if os.path.dirname(dest_path) else ".", exist_ok=True)
            wb.save(dest_path)
            
            logger.info(f"Exported summary: {len(invoices)} rows → {dest_path}")
            return {
                "success": True,
                "file_path": dest_path,
                "row_count": len(invoices),
                "error_msg": "",
            }
        except Exception as e:
            logger.error(f"Export summary error: {e}")
            return {"success": False, "file_path": "", "row_count": 0, "error_msg": str(e)}
    
    # ═══════════════════════════════════════════════════════
    # EXPORT DETAIL (Summary + Chi tiết hàng hóa)
    # ═══════════════════════════════════════════════════════
    
    @staticmethod
    def export_detail(
        invoices: List[InvoiceData],
        dest_path: str,
        summary_columns: List[ColumnConfig],
        detail_columns: List[ColumnConfig],
    ) -> dict:
        """Export 2 sheets: Sheet 1 = Summary, Sheet 2 = Detail (mỗi dòng hàng hóa).
        
        Returns:
            {"success": bool, "file_path": str, "row_count": int, "error_msg": str}
        """
        try:
            wb = Workbook()
            
            # ── Sheet 1: Summary ──
            ws1 = wb.active
            ws1.title = "Tổng hợp"
            ExcelExportService._write_sheet_data(ws1, invoices, summary_columns, "summary")
            
            # ── Sheet 2: Detail ──
            ws2 = wb.create_sheet("Chi tiết hàng hóa")
            
            # Headers cho detail: thêm cột HĐ tham chiếu (Ký hiệu + Số HĐ + Ngày lập)
            ref_cols = [
                ColumnConfig(column_key="_ky_hieu", display_name="Ký hiệu HĐ", format_type="text", width=100),
                ColumnConfig(column_key="_so_hd", display_name="Số HĐ", format_type="text", width=80),
                ColumnConfig(column_key="_ngay_lap", display_name="Ngày lập", format_type="date", width=100),
                ColumnConfig(column_key="_ten_ban", display_name="Tên NB", format_type="text", width=200),
            ]
            all_detail_cols = ref_cols + detail_columns
            
            # Title
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_detail_cols))
            title_cell = ws2.cell(row=1, column=1, value="CHI TIẾT HÀNG HÓA / DỊCH VỤ")
            title_cell.font = TITLE_FONT
            title_cell.alignment = Alignment(horizontal="center")
            
            # Headers
            header_row = 3
            for col_idx, col_cfg in enumerate(all_detail_cols, 1):
                cell = ws2.cell(row=header_row, column=col_idx, value=col_cfg.display_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
                ws2.column_dimensions[get_column_letter(col_idx)].width = max(
                    col_cfg.width / 8, 12
                )
            
            # Data: mỗi dòng hàng hóa từ mỗi HĐ
            data_row = header_row + 1
            for inv in invoices:
                for item in inv.hang_hoa:
                    col_idx = 1
                    
                    # Reference columns
                    for ref in ref_cols:
                        if ref.column_key == "_ky_hieu":
                            val = inv.ky_hieu
                        elif ref.column_key == "_so_hd":
                            val = inv.so_hd
                        elif ref.column_key == "_ngay_lap":
                            val = inv.ngay_lap
                        elif ref.column_key == "_ten_ban":
                            val = inv.ten_ban
                        else:
                            val = ""
                        cell = ws2.cell(row=data_row, column=col_idx, value=val)
                        cell.font = BODY_FONT
                        cell.border = THIN_BORDER
                        ExcelExportService._apply_format(cell, ref.format_type)
                        col_idx += 1
                    
                    # Detail columns
                    for col_cfg in detail_columns:
                        value = ExcelExportService._get_detail_value(item, col_cfg)
                        cell = ws2.cell(row=data_row, column=col_idx, value=value)
                        cell.font = BODY_FONT
                        cell.border = THIN_BORDER
                        ExcelExportService._apply_format(cell, col_cfg.format_type)
                        col_idx += 1
                    
                    data_row += 1
            
            # Auto-filter cho sheet 2
            total_detail_rows = data_row - header_row - 1
            if total_detail_rows > 0:
                ws2.auto_filter.ref = f"A{header_row}:{get_column_letter(len(all_detail_cols))}{data_row - 1}"
            
            # Save
            os.makedirs(os.path.dirname(dest_path) if os.path.dirname(dest_path) else ".", exist_ok=True)
            wb.save(dest_path)
            
            logger.info(f"Exported detail: {len(invoices)} HĐ, {total_detail_rows} dòng hàng → {dest_path}")
            return {
                "success": True,
                "file_path": dest_path,
                "row_count": total_detail_rows,
                "error_msg": "",
            }
        except Exception as e:
            logger.error(f"Export detail error: {e}")
            return {"success": False, "file_path": "", "row_count": 0, "error_msg": str(e)}
    
    # ═══════════════════════════════════════════════════════
    # HELPERS
    # ═══════════════════════════════════════════════════════
    
    @staticmethod
    def _write_sheet_data(ws, invoices, columns, mode):
        """Ghi sheet summary."""
        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value="DANH SÁCH HÓA ĐƠN ĐIỆN TỬ")
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")
        
        header_row = 3
        for col_idx, col_cfg in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_cfg.display_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = max(col_cfg.width / 8, 12)
        
        for row_idx, inv in enumerate(invoices, header_row + 1):
            for col_idx, col_cfg in enumerate(columns, 1):
                value = ExcelExportService._get_summary_value(inv, col_cfg, row_idx - header_row)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                ExcelExportService._apply_format(cell, col_cfg.format_type)
        
        if invoices:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(invoices)}"
    
    @staticmethod
    def _get_summary_value(inv: InvoiceData, col: ColumnConfig, row_num: int):
        """Lấy giá trị cho 1 ô trong bảng summary."""
        key = col.column_key
        
        # Cột STT (tự động đánh số)
        if key == "stt":
            return row_num
        
        # Cột Tầng 1 (fixed)
        field_map = {
            "ky_hieu": inv.ky_hieu,
            "so_hd": inv.so_hd,
            "ngay_lap": inv.ngay_lap,
            "mst_ban": inv.mst_ban,
            "ten_ban": inv.ten_ban,
            "mst_mua": inv.mst_mua,
            "ten_mua": inv.ten_mua,
            "tong_chua_thue": inv.tong_chua_thue,
            "tong_thue": inv.tong_thue,
            "tong_thanh_toan": inv.tong_thanh_toan_so,
            "trang_thai": inv.status_label,
            "vendor": inv.nha_cung_cap,
            "mccqt": inv.mccqt,
            "fkey": inv.fkey,
            "portal_link": inv.portal_link,
        }
        
        if key in field_map:
            value = field_map[key]
            # Chuyển số dạng string sang float cho các cột currency/number
            if col.format_type in ("currency", "number"):
                return ExcelExportService._to_number(value)
            return value
        
        # Cột Tầng 2 (dynamic extras)
        if col.is_dynamic:
            scope, field_key = parse_dynamic_column_key(key)
            extras_map = {
                "header": inv.extras_header,
                "seller": inv.extras_seller,
                "buyer": inv.extras_buyer,
                "payment": inv.extras_payment,
                "invoice": inv.extras_invoice,
            }
            extras = extras_map.get(scope, {})
            value = extras.get(field_key, "")
            if col.format_type in ("currency", "number"):
                return ExcelExportService._to_number(value)
            return value
        
        return ""
    
    @staticmethod
    def _get_detail_value(item, col: ColumnConfig):
        """Lấy giá trị cho 1 ô trong bảng chi tiết hàng hóa."""
        key = col.column_key
        
        field_map = {
            "stt": item.stt,
            "tinh_chat": item.tinh_chat,
            "ma_hang": item.ma_hang,
            "ten_hang": item.ten_hang,
            "don_vi_tinh": item.don_vi_tinh,
            "so_luong": item.so_luong,
            "don_gia": item.don_gia,
            "ty_le_ck": item.ty_le_ck,
            "so_tien_ck": item.so_tien_ck,
            "thue_suat": item.thue_suat,
            "thanh_tien": item.thanh_tien,
        }
        
        if key in field_map:
            value = field_map[key]
            if col.format_type in ("currency", "number"):
                return ExcelExportService._to_number(value)
            return value
        
        # Tầng 2: extras của dòng hàng
        if col.is_dynamic:
            _, field_key = parse_dynamic_column_key(key)
            value = item.extras.get(field_key, "")
            if col.format_type in ("currency", "number"):
                return ExcelExportService._to_number(value)
            return value
        
        return ""
    
    @staticmethod
    def _apply_format(cell, format_type: str):
        """Áp dụng format theo loại cột."""
        if format_type in ("currency", "number"):
            cell.number_format = VND_FORMAT
            cell.alignment = BODY_ALIGN_RIGHT
        elif format_type == "date":
            cell.alignment = BODY_ALIGN_CENTER
        else:
            cell.alignment = BODY_ALIGN_LEFT
    
    @staticmethod
    def _to_number(value) -> Optional[float]:
        """Chuyển string sang number (trả None nếu rỗng)."""
        if not value or str(value).strip() == "":
            return None
        try:
            return float(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return None
